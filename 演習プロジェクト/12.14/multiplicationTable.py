import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb['Sheet']

num = int(input('command line :'))

for i in range(1, num+1) :

    sheet.cell(row=(i+1), column=1).value = i
    sheet.cell(row=(i+1), column=1).font = Font(bold=True)
    sheet.cell(row=1, column=(i+1)).value = i
    sheet.cell(row=1, column=(i+1)).font = Font(bold=True)

    for j in range(1, num+1) :
        
        sheet.cell(row=(i+1), column=(j+1)).value = i*j

wb.save('multiplicationTable.xlsx')
