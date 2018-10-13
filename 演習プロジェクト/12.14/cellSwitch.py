import openpyxl

line = []

wb = openpyxl.load_workbook('cell_change_before.xlsx')
sheet = wb['Sheet1']

for i in range(1, sheet.max_column+1) :
    temp = []
    for j in range(1, sheet.max_row+1) :
        temp.append(sheet.cell(row=j, column=i).value)
        sheet.cell(row=j, column=i).value = ''
    line.append(temp)


for i in range(0, len(line)) :
    for j in range(0,len(line[i])) :
        
        sheet.cell(row=i+1, column=j+1).value = line[i][j]

wb.save('cell_change_after.xlsx')
