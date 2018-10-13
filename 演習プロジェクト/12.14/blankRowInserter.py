import openpyxl

def chkInt(num) :

     while True :
        if num > 0 :
            break
        print('1以上の整数を入力してください。')
        num = int(input())

     return num

wb = openpyxl.load_workbook('myProduceBefore.xlsx')
sheet = wb['Sheet1']

N = chkInt(int(input('第1の整数 : ')))
M = chkInt(int(input('第2の整数 : ')))

for i in range(M) :
    for j in range(1, sheet.max_column+1) :

        sheet.cell(row=(N+i), column=j).value = ''

wb.save('myProduceAfter.xlsx')
